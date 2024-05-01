import openpyxl

file=r"F:\Python & Selenium by pavan sir\SeleniumDDT Practise notes\writefile.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook['Sheet1']
# sheet=workbook.active               #  -> There is wrtiing another way

# #1)--For Sending the same type of Data
#--------------------------------------
for r in range(1,6):
    for c in range(1,4):
        sheet.cell(r,c).value="welcome2"
    workbook.save(file)

#2)-For Sending the diffrent type of Data
#-------------------------------
sheet.cell(1,1).value=123
sheet.cell(1,2).value="Smith"
sheet.cell(1,3).value="Engineer"


sheet.cell(2,1).value=456
sheet.cell(2,2).value="John"
sheet.cell(2,3).value="Manager"

sheet.cell(3,1).value=567
sheet.cell(3,2).value="David"
sheet.cell(3,3).value="Developer"
workbook.save(file)                         # save the file after enterign the data
