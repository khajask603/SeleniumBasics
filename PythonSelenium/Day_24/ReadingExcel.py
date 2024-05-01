import openpyxl


file=r"F:\Python & Selenium by pavan sir\SeleniumDDT Practise notes\data.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook["Sheet1"]

rows=sheet.max_row
colu=sheet.max_column

#Reading all the rows & columns from excel sheet
for r in range(1,rows+1):
    for c in range(1,colu+1):
        print(sheet.cell(r,c).value, end='   ||     ')
    print()